import deepl
import openpyxl as xl

# Copy the name of the Excel file here
file_name = "#.xlsx"

# You can get a free authentication key here https://support.deepl.com/hc/en-us/articles/360020695820-Authentication-Key
authentication_key = "################################"
translator = deepl.Translator(authentication_key)

# Place the Excel file in the same folder
fhandle = xl.load_workbook(f"{file_name}")
sheet = fhandle.active
tmp = open("input.txt", 'w')
tt = open("output.txt", 'w')
input_path = "input.txt"
output_path = "output.txt"


def read_text_chunks(output_path):
    with open(output_path, 'r', encoding='utf-8') as file:
        text = file.read()
    return text.split('\n\n')  # Assuming chunks are separated by blank lines


def update_excel(file_name, text_chunks):
    workbook = xl.load_workbook(file_name)
    active_sheet = workbook.active

    for row_index in range(1, active_sheet.max_row + 1):
        cell_value = active_sheet.cell(row=row_index, column=1).value

        if cell_value:
            text_chunk = text_chunks.pop(0) if text_chunks else ""
            active_sheet.cell(row=row_index, column=2, value=text_chunk)
            
    workbook.save(f"Translated {file_name}")
    workbook.close()

for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    if cell.value is not None:
        tbt = cell.value
        str(tbt)
        tmp.write(tbt + '\n\n')

fhandle.close()
tmp.close()
translator.translate_document_from_filepath(input_path, output_path, target_lang='EN-US')
text_chunks = read_text_chunks(output_path)
update_excel(file_name, text_chunks)
